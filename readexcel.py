# An example of creating a workbook using openpyxl
# excel-python.py
from openpyxl import Workbook
from time import strftime
Wb_test = Workbook("/YOUR/FILE/PATH.xlsx")

# Select the active worksheet
ws = Wb_test.active

no = '1'

time = strftime("%H:%M:%S")
print("time:", time)

date_time = strftime("%m/%d/%Y")
print("date:",date_time)

ws['A1'] = no
ws['B1'] = date_time
ws['C1'] = time
Wb_test.save("/YOUR/FILE/PATH.xlsx")
