# import openpyxl module 
from openpyxl import load_workbook  
  
# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
arr_str = "Num","Date","Time","T/No","G.w(kg)","T.w(kg)","N.w(kg)"
wb = load_workbook('/YOUR/FILE/PATH.xlsx')  
sheet = wb.active
 
print(arr_str[1])
for x in range(1, 8):
    sheet.cell(row=3, column=x).value = arr_str[x-1]
wb.save('/YOUR/FILE/PATH.xlsx')  